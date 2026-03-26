from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from phoneme_psychopy.schedule_loader import load_trials_from_workbook


class ScheduleLoaderTests(unittest.TestCase):
    def test_load_trials_from_workbook_uses_configured_playback_track_order(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "schedule.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Template"

            worksheet.cell(row=3, column=2, value="Track")
            worksheet.cell(row=3, column=3, value="SNR")
            worksheet.cell(row=3, column=4, value="07:00")

            track_rows = [
                ("1A", 10.0, "a"),
                ("1B", 5.0, "b"),
                ("1C", 0.0, "c"),
                ("1D", -5.0, "d"),
                ("1E", -10.0, "e"),
                ("2A", 10.0, "f"),
                ("2B", 5.0, "g"),
                ("2C", 0.0, "h"),
                ("2D", -5.0, "i"),
                ("2E", -10.0, "j"),
            ]

            current_row = 4
            for track_id, snr, phoneme in track_rows:
                worksheet.cell(row=current_row, column=2, value=track_id)
                worksheet.cell(row=current_row, column=3, value=snr)
                worksheet.cell(row=current_row, column=4, value=phoneme)
                current_row += 1

            workbook.save(workbook_path)

            trials = load_trials_from_workbook(workbook_path)

            self.assertEqual(
                [trial.track_id for trial in trials],
                ["1C", "2B", "1E", "2D", "2A", "1D", "2E", "1A", "2C", "1B"],
            )
            self.assertEqual([trial.trial_index for trial in trials], list(range(1, 11)))


if __name__ == "__main__":
    unittest.main()
