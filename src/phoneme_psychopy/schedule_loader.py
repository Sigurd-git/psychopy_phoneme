from __future__ import annotations

from dataclasses import asdict
from datetime import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .models import TrialDefinition


TIME_COLUMNS_START_INDEX_1_BASED = 4
TIME_COLUMNS_END_INDEX_1_BASED = 23
TRACK_LABEL_PREFIXES = tuple(
    f"{noise_index}{snr_band}"
    for noise_index in (1, 2)
    for snr_band in ("A", "B", "C", "D", "E")
)


def load_trials_from_workbook(schedule_path: Path, sheet_name: str = "Template") -> list[TrialDefinition]:
    """Load the stimulus schedule workbook and convert it into a flat trial table.

    The current workbook stores trial content in a wide layout where each track row contains
    phoneme values spread across many onset columns. This parser converts each populated cell
    into one trial record while preserving workbook provenance for traceability.
    """

    workbook = load_workbook(schedule_path, data_only=True)
    worksheet = workbook[sheet_name]

    onset_row_index = find_row_index_containing_value(worksheet, "Track")
    if onset_row_index is None:
        raise ValueError("Could not find the workbook row that defines the onset columns.")

    onset_labels = [
        format_onset_label(worksheet.cell(row=onset_row_index, column=column_index).value)
        for column_index in range(TIME_COLUMNS_START_INDEX_1_BASED, TIME_COLUMNS_END_INDEX_1_BASED + 1)
    ]

    parsed_trials: list[TrialDefinition] = []
    for row_index in range(onset_row_index + 1, worksheet.max_row + 1):
        track_value = worksheet.cell(row=row_index, column=2).value
        if not isinstance(track_value, str) or track_value not in TRACK_LABEL_PREFIXES:
            continue

        snr_value = worksheet.cell(row=row_index, column=3).value
        for onset_position, column_index in enumerate(
            range(TIME_COLUMNS_START_INDEX_1_BASED, TIME_COLUMNS_END_INDEX_1_BASED + 1)
        ):
            phoneme_value = worksheet.cell(row=row_index, column=column_index).value
            if phoneme_value in (None, ""):
                continue

            session_type = infer_session_type_from_track(track_value)
            trial_definition = TrialDefinition(
                track_id=track_value,
                snr=float(snr_value),
                onset_label=onset_labels[onset_position],
                phoneme=str(phoneme_value),
                session_type=session_type,
                trial_index=len(parsed_trials) + 1,
                source_sheet=sheet_name,
                source_row=row_index,
                source_column=excel_column_name(column_index),
            )
            parsed_trials.append(trial_definition)

    return parsed_trials


def find_row_index_containing_value(worksheet, target_value: str) -> int | None:
    """Return the first worksheet row index whose second column matches the target value."""

    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=2).value == target_value:
            return row_index
    return None


def infer_session_type_from_track(track_id: str) -> str:
    """Map workbook track labels to the correct noise family.

    The workbook groups white-noise tracks under the `1X` family and babble-noise tracks under the
    `2X` family, where X indicates the SNR band letter.
    """

    if track_id.startswith("1"):
        return "white"
    if track_id.startswith("2"):
        return "babble"
    raise ValueError(f"Unrecognized track_id for session inference: {track_id}")


def build_trial_preview_table(trials: list[TrialDefinition]) -> pd.DataFrame:
    """Return a compact preview table for debugging and validation."""

    return pd.DataFrame([asdict(trial) for trial in trials])


def format_onset_label(cell_value: object) -> str:
    """Convert workbook onset cells into readable labels."""

    if isinstance(cell_value, time):
        return cell_value.strftime("%M:%S")
    if cell_value is None:
        return ""
    return str(cell_value)


def excel_column_name(column_index_1_based: int) -> str:
    """Convert a one-based numeric index into an Excel column label."""

    label = ""
    column_index = column_index_1_based
    while column_index > 0:
        column_index, remainder = divmod(column_index - 1, 26)
        label = chr(65 + remainder) + label
    return label
